"""Excel файлов (.xlsx и .xls) парсер.

PODDERZHIVAYET:
- .xlsx (Excel 2007+) - openpyxl
- .xls (Excel 97-2003) - pandas (NO xlrd)

EXTRACT:
- All sheets to plain text
- Table structure preserved
- Empty cells handled
"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parse Excel files (.xlsx and .xls)."""

    def extract_text(self, file_path: Path) -> str:
        """Парсить Excel файл.
        
        Args:
            file_path: Path to Excel file (.xlsx or .xls)
            
        Returns:
            Extracted text from all sheets
            
        Raises:
            ValueError: If file format is unsupported or corrupted
        """
        file_name = file_path.name.lower()
        
        # Determine format from extension
        if file_name.endswith('.xlsx'):
            logger.info(f"Parsing .xlsx file: {file_path.name}")
            return self._parse_xlsx(file_path)
        
        elif file_name.endswith('.xls'):
            logger.info(f"Parsing .xls file: {file_path.name}")
            return self._parse_xls(file_path)
        
        else:
            raise ValueError(f"Unsupported Excel format: {file_name}")

    def _parse_xlsx(self, file_path: Path) -> str:
        """Парсить .xlsx файл используя openpyxl."""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path, data_only=True)
            logger.info(f"XLSX: Loaded workbook with {len(workbook.sheetnames)} sheets")
            
            all_text = []
            
            for sheet_name in workbook.sheetnames:
                logger.info(f"XLSX: Processing sheet '{sheet_name}'")
                sheet = workbook[sheet_name]
                
                # Add sheet name
                all_text.append(f"\n=== {sheet_name} ===")
                
                # Extract rows as tab-separated values
                for row in sheet.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if not any(cell is not None for cell in row):
                        continue
                    
                    # Convert cells to strings, handle None
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    line = "\t".join(cells)
                    all_text.append(line)
            
            workbook.close()
            text = "\n".join(all_text).strip()
            logger.info(f"XLSX: Extracted {len(text)} characters")
            return text
        
        except ImportError:
            logger.error("openpyxl not installed")
            raise ValueError("openpyxl library required for .xlsx files")
        
        except Exception as e:
            logger.error(f"XLSX parse error: {e}")
            raise ValueError(f"Failed to parse .xlsx file: {str(e)}")

    def _parse_xls(self, file_path: Path) -> str:
        """Парсить .xls файл используя pandas (WITHOUT xlrd)."""
        try:
            import pandas as pd
            
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            logger.info(f"XLS: Loaded workbook with {len(excel_file.sheet_names)} sheets")
            
            all_text = []
            
            for sheet_name in excel_file.sheet_names:
                logger.info(f"XLS: Processing sheet '{sheet_name}'")
                
                try:
                    # Read sheet - pandas handles both .xls and .xlsx
                    df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
                    
                    # Skip completely empty sheets
                    if df.empty:
                        logger.info(f"XLS: Sheet '{sheet_name}' is empty, skipping")
                        continue
                    
                    # Add sheet name
                    all_text.append(f"\n=== {sheet_name} ===")
                    
                    # Add headers
                    headers = df.columns.tolist()
                    all_text.append("\t".join(str(h) for h in headers))
                    
                    # Add rows
                    for _, row in df.iterrows():
                        # Handle NaN values
                        cells = [str(val) if pd.notna(val) else "" for val in row]
                        line = "\t".join(cells)
                        all_text.append(line)
                
                except Exception as sheet_error:
                    logger.warning(f"XLS: Error reading sheet '{sheet_name}': {sheet_error}")
                    continue
            
            text = "\n".join(all_text).strip()
            logger.info(f"XLS: Extracted {len(text)} characters")
            return text
        
        except ImportError:
            logger.error("pandas not installed for .xls support")
            raise ValueError(
                "pandas library required for .xls files. "
                "Install with: pip install pandas"
            )
        
        except Exception as e:
            logger.error(f"XLS parse error: {e}")
            raise ValueError(f"Failed to parse .xls file: {str(e)}")
