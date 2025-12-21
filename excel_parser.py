"""Excel file parser for .xlsx and .xls formats.

Extracts text from Excel spreadsheets using openpyxl (.xlsx) and xlrd (.xls).
Handles multiple sheets and cell values.

Fixes 2025-12-20 23:27:
- Добавлена поддержка старых .xls файлов через xlrd
- Автоматическое определение формата по расширению
- Graceful fallback если библиотека отсутствует
"""

import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parser for extracting text from Excel files.
    
    Supports: .xlsx (openpyxl), .xls (xlrd)
    Extracts text from all sheets and cells.
    """
    
    def extract_text(self, file_path: Path) -> str:
        """Extract text from Excel file.
        
        Reads all sheets and combines text from cells.
        Skips empty cells and organizes by sheet.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            str: Extracted text from all sheets
            
        Raises:
            ImportError: If required library is not installed
            Exception: If file cannot be parsed
        """
        file_ext = file_path.suffix.lower()
        
        if file_ext == '.xlsx':
            return self._extract_xlsx(file_path)
        elif file_ext == '.xls':
            return self._extract_xls(file_path)
        else:
            raise ValueError(f"Unsupported Excel format: {file_ext}")
    
    def _extract_xlsx(self, file_path: Path) -> str:
        """Extract text from .xlsx file using openpyxl.
        
        Args:
            file_path: Path to .xlsx file
            
        Returns:
            str: Extracted text
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl is required for .xlsx file support")
        
        try:
            logger.info(f"Extracting text from Excel (.xlsx): {file_path.name}")
            
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            all_text: list[str] = []
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                
                sheet_text: list[str] = []
                sheet_text.append(f"=== Sheet: {sheet_name} ===")
                
                # Extract text from cells
                for row in sheet.iter_rows(values_only=True):
                    row_values = []
                    for cell_value in row:
                        if cell_value is not None:
                            row_values.append(str(cell_value).strip())
                    
                    if row_values:  # Skip empty rows
                        sheet_text.append(" | ".join(row_values))
                
                if len(sheet_text) > 1:  # Only add if sheet has content
                    all_text.append("\n".join(sheet_text))
            
            workbook.close()
            
            result = "\n\n".join(all_text)
            logger.info(f"Successfully extracted {len(result)} chars from .xlsx")
            return result
        
        except Exception as e:
            logger.error(f"Error parsing .xlsx file {file_path.name}: {e}")
            raise
    
    def _extract_xls(self, file_path: Path) -> str:
        """Extract text from old .xls file using xlrd.
        
        Args:
            file_path: Path to .xls file
            
        Returns:
            str: Extracted text
        """
        try:
            import xlrd
        except ImportError:
            logger.error("xlrd not installed. Install with: pip install xlrd")
            raise ImportError(
                "xlrd is required for old .xls file support. "
                "Install with: pip install xlrd"
            )
        
        try:
            logger.info(f"Extracting text from Excel (.xls): {file_path.name}")
            
            # Open workbook
            workbook = xlrd.open_workbook(file_path)
            all_text: list[str] = []
            
            # Process each sheet
            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                sheet_name = sheet.name
                logger.info(f"Processing sheet: {sheet_name}")
                
                sheet_text: list[str] = []
                sheet_text.append(f"=== Sheet: {sheet_name} ===")
                
                # Extract text from cells
                for row_idx in range(sheet.nrows):
                    row_values = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        # Skip empty cells
                        if cell_value is not None and str(cell_value).strip():
                            row_values.append(str(cell_value).strip())
                    
                    if row_values:  # Skip empty rows
                        sheet_text.append(" | ".join(row_values))
                
                if len(sheet_text) > 1:  # Only add if sheet has content
                    all_text.append("\n".join(sheet_text))
            
            result = "\n\n".join(all_text)
            logger.info(f"Successfully extracted {len(result)} chars from .xls")
            return result
        
        except Exception as e:
            logger.error(f"Error parsing .xls file {file_path.name}: {e}")
            raise
